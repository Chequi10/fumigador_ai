from openpyxl import Workbook, load_workbook
import threading
import time

lock = threading.Lock()
archivo_excel = "/home/ezequiel/Dropbox/Ezequiel/Bayer_solutions/ia/yolov5/seguimiento.xlsx"

def iniciar_archivo_excel():
    """Carga o crea el archivo Excel."""
    global libro, hoja
    try:
        # Si el archivo ya existe, lo cargamos
        libro = load_workbook(archivo_excel)
        hoja = libro.active
    except FileNotFoundError:
        # Si el archivo no existe, creamos uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Tramas GPS"
        hoja.append(["Fecha", "Hora", "Latitud", "Longitud", "Velocidad (km/h)"])  # Encabezados
datos_a_guardar = []

def guardar_datos_en_excel():
    """Guarda los datos de la lista en el archivo Excel"""
    global libro, hoja, datos_a_guardar
    try:
        with lock:
            for fila in datos_a_guardar:
                hoja.append(fila)
            libro.save(archivo_excel)
            print(f"Datos guardados en {archivo_excel}")
    except Exception as e:
        print(f"Error al guardar datos en Excel: {e}")


